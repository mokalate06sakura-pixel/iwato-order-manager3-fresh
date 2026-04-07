import re
from pathlib import Path
from typing import Dict, Tuple, List

import pandas as pd
import openpyxl
from openpyxl.worksheet.worksheet import Worksheet


def find_col_by_keywords(df: pd.DataFrame, keywords: list[str]) -> str:
    cols = [str(c) for c in df.columns]
    for c in cols:
        if all(k in c for k in keywords):
            return c
    raise KeyError(
        f"列が見つかりません: keywords={keywords}\n利用可能な列:\n" + "\n".join(cols)
    )


INVALID_SHEET_CHARS = r'[:\\/*?\[\]]'

HEADER_CELL_FACILITY = "I2"
TOKUYOU_LABEL = "特養いわと"
YUHOUSE_LABEL = "ユーハウスいわと"

COL_SUPPLIER = "仕入先"
COL_USE_DATE = "使用日"
COL_FOOD_NAME = "食品名"
COL_SPEC = "換算値"

SUPPLIER_NAME = "丸八ヒロタ"

# テンプレの実シート名
TEMPLATE_SHEET_NAME_TOKUYOU = "丸八ヒロタ発注書(介護老人福祉施設いわと）"
TEMPLATE_SHEET_NAME_YUHOUSE = "丸八ヒロタ発注書(ユーハウス）"

TAG_SHEET_NAME = "タグ"

FIXED_FIRST_ROW = 6
APPEND_START_ROW = 22
APPEND_MAX_ROWS = 7

COL_OUT_USE_DATE = 1
COL_OUT_CODE = 2
COL_OUT_NAME_2 = 4
COL_OUT_SPEC = 5
COL_OUT_RESIDENT = 6
COL_OUT_STAFF = 7


def sanitize_sheet_title(title: str, existing: set[str]) -> str:
    t = str(title)
    t = re.sub(INVALID_SHEET_CHARS, "-", t)
    t = t.strip()

    if t.startswith("'"):
        t = t[1:]
    if t.endswith("'"):
        t = t[:-1]
    if not t:
        t = "Sheet"

    t = t[:31]

    base = t
    i = 2
    while t in existing:
        suffix = f"_{i}"
        t = base[: 31 - len(suffix)] + suffix
        i += 1
    return t


def _norm(s: object) -> str:
    if s is None:
        return ""
    t = str(s)
    t = t.replace("\u3000", " ")
    t = re.sub(r"\s+", " ", t)
    return t.strip()


def load_tag_mapping(tag_xlsm_path: str | Path) -> Dict[str, Tuple[str, str, str]]:
    wb = openpyxl.load_workbook(Path(tag_xlsm_path), data_only=True, keep_vba=True)
    ws = wb[TAG_SHEET_NAME]

    mapping: Dict[str, Tuple[str, str, str]] = {}
    for r in range(2, ws.max_row + 1):
        code = ws.cell(r, 1).value
        maru_name = ws.cell(r, 2).value
        spec = ws.cell(r, 3).value
        heart_name = ws.cell(r, 4).value

        k = _norm(heart_name)
        if not k or code is None:
            continue

        mapping[k] = (str(code), str(maru_name or ""), str(spec or ""))

    return mapping


def _read_kenshu(kenshu_xlsx_path: str | Path) -> pd.DataFrame:
    df = pd.read_excel(Path(kenshu_xlsx_path))
    df = df[df[COL_SUPPLIER].astype(str) == SUPPLIER_NAME].copy()
    df = df[df[COL_USE_DATE].notna()].copy()
    return df


def _build_fixed_row_index(ws: Worksheet) -> Dict[str, int]:
    idx: Dict[str, int] = {}
    r = FIXED_FIRST_ROW
    while r < APPEND_START_ROW:
        code = ws.cell(r, COL_OUT_CODE).value
        if code is None or str(code).strip() == "":
            break
        idx[str(code).strip()] = r
        r += 1
    return idx


def _clear_sheet_quantities(ws: Worksheet) -> None:
    r = FIXED_FIRST_ROW
    while r < APPEND_START_ROW:
        code = ws.cell(r, COL_OUT_CODE).value
        if code is None or str(code).strip() == "":
            break
        ws.cell(r, COL_OUT_USE_DATE).value = None
        ws.cell(r, COL_OUT_RESIDENT).value = None
        ws.cell(r, COL_OUT_STAFF).value = None
        r += 1

    for rr in range(APPEND_START_ROW, APPEND_START_ROW + APPEND_MAX_ROWS):
        ws.cell(rr, COL_OUT_USE_DATE).value = None
        ws.cell(rr, COL_OUT_CODE).value = None
        ws.cell(rr, COL_OUT_NAME_2).value = None
        ws.cell(rr, COL_OUT_SPEC).value = None
        ws.cell(rr, COL_OUT_RESIDENT).value = None
        ws.cell(rr, COL_OUT_STAFF).value = None


def _write_append_row(
    ws: Worksheet,
    rr: int,
    use_date: object,
    food_name: str,
    spec: str,
    qty_resident: float,
    qty_staff: float,
) -> None:
    ws.cell(rr, COL_OUT_USE_DATE).value = use_date
    ws.cell(rr, COL_OUT_NAME_2).value = food_name
    ws.cell(rr, COL_OUT_SPEC).value = spec
    ws.cell(rr, COL_OUT_RESIDENT).value = qty_resident if qty_resident != 0 else None
    ws.cell(rr, COL_OUT_STAFF).value = qty_staff if qty_staff != 0 else None


def _copy_base_sheet(wb, base_ws, title, facility_mode: str):
    ws2 = wb.copy_worksheet(base_ws)

    existing = set(wb.sheetnames)
    safe_title = sanitize_sheet_title(title, existing)
    ws2.title = safe_title

    _clear_sheet_quantities(ws2)

    if facility_mode == "yuhouse":
        ws2[HEADER_CELL_FACILITY] = YUHOUSE_LABEL
    else:
        ws2[HEADER_CELL_FACILITY] = TOKUYOU_LABEL

    return ws2


def generate_maruhachi_order_workbook(
    kenshu_xlsx_path: str | Path,
    template_xlsm_path: str | Path,
    tag_xlsm_path: str | Path,
    facility_mode: str,
    out_path: str | Path,
) -> Path:
    if facility_mode not in ("tokuyou", "yuhouse"):
        raise ValueError("facility_mode must be 'tokuyou' or 'yuhouse'")

    kenshu_xlsx_path = Path(kenshu_xlsx_path)
    template_xlsm_path = Path(template_xlsm_path)
    tag_xlsm_path = Path(tag_xlsm_path)
    out_path = Path(out_path)

    df = _read_kenshu(kenshu_xlsx_path)

    if facility_mode == "tokuyou":
        col_res = find_col_by_keywords(df, ["介護老人福祉施設いわと", "入所者"])
        col_staff = find_col_by_keywords(df, ["介護老人福祉施設いわと", "職員"])
    else:
        try:
            col_res = find_col_by_keywords(df, ["ケアハウス", "入所者"])
        except KeyError:
            col_res = find_col_by_keywords(df, ["ユーハウス", "入所者"])
        col_staff = None

    tag_map = load_tag_mapping(tag_xlsm_path)

    wb = openpyxl.load_workbook(template_xlsm_path, keep_vba=True)

    if facility_mode == "tokuyou":
        base_ws = wb[TEMPLATE_SHEET_NAME_TOKUYOU]
    else:
        base_ws = wb[TEMPLATE_SHEET_NAME_YUHOUSE]

    fixed_row_index = _build_fixed_row_index(base_ws)
    _clear_sheet_quantities(base_ws)

    use_dates = sorted(df[COL_USE_DATE].dropna().astype(str).unique().tolist())

    created_sheets = []

    for use_date in use_dates:
        ddf = df[df[COL_USE_DATE].astype(str) == use_date].copy()
        ddf[col_res] = pd.to_numeric(ddf[col_res], errors="coerce").fillna(0)

        if col_staff is not None:
            ddf[col_staff] = pd.to_numeric(ddf[col_staff], errors="coerce").fillna(0)
            col_staff_tmp = col_staff
        else:
            ddf["_staff"] = 0
            col_staff_tmp = "_staff"

        grouped = (
            ddf.groupby([COL_FOOD_NAME, COL_SPEC], dropna=False)[[col_res, col_staff_tmp]]
            .sum()
            .reset_index()
        )

        sheet_title = str(use_date)
        ws = _copy_base_sheet(wb, base_ws, sheet_title, facility_mode)
        created_sheets.append(ws.title)

        append_items: List[Tuple[str, str, float, float]] = []

        for _, row in grouped.iterrows():
            food_name = _norm(row[COL_FOOD_NAME])
            spec = str(row[COL_SPEC] or "")
            qty_res = float(row[col_res] or 0)
            qty_staff = float(row[col_staff_tmp] or 0)

            if qty_res == 0 and qty_staff == 0:
                continue

            tag_key = _norm(food_name)
            if tag_key in tag_map:
                code, maru_name, maru_spec = tag_map[tag_key]
                if code in fixed_row_index:
                    rr = fixed_row_index[code]
                    ws.cell(rr, COL_OUT_USE_DATE).value = use_date
                    ws.cell(rr, COL_OUT_RESIDENT).value = qty_res if qty_res != 0 else None
                    ws.cell(rr, COL_OUT_STAFF).value = qty_staff if qty_staff != 0 else None
                else:
                    append_items.append((food_name, spec, qty_res, qty_staff))
            else:
                append_items.append((food_name, spec, qty_res, qty_staff))

        if append_items:
            page = 1
            pos = 0
            while pos < len(append_items):
                if page == 1:
                    cur_ws = ws
                else:
                    cur_ws = _copy_base_sheet(
                        wb,
                        base_ws,
                        f"{use_date}_{page}ページ目",
                        facility_mode,
                    )
                    created_sheets.append(cur_ws.title)

                for rr in range(APPEND_START_ROW, APPEND_START_ROW + APPEND_MAX_ROWS):
                    if pos >= len(append_items):
                        break

                    f, sp, qr, qs = append_items[pos]
                    _write_append_row(cur_ws, rr, use_date, f, sp, qr, qs)
                    pos += 1

                page += 1


    out_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out_path)
    return out_path


def generate_maruhachi_order_forms_both_facilities(
    kenshu_xlsx_path: str | Path,
    template_xlsm_path: str | Path,
    tag_xlsm_path: str | Path,
    out_dir: str | Path,
    out_prefix: str = "丸八発注書",
) -> Tuple[Path, Path]:
    out_dir = Path(out_dir)
    out_dir.mkdir(parents=True, exist_ok=True)

    tokuyou_path = out_dir / f"{out_prefix}_特養.xlsm"
    yuhouse_path = out_dir / f"{out_prefix}_ユーハウス.xlsm"

    p1 = generate_maruhachi_order_workbook(
        kenshu_xlsx_path=kenshu_xlsx_path,
        template_xlsm_path=template_xlsm_path,
        tag_xlsm_path=tag_xlsm_path,
        facility_mode="tokuyou",
        out_path=tokuyou_path,
    )
    p2 = generate_maruhachi_order_workbook(
        kenshu_xlsx_path=kenshu_xlsx_path,
        template_xlsm_path=template_xlsm_path,
        tag_xlsm_path=tag_xlsm_path,
        facility_mode="yuhouse",
        out_path=yuhouse_path,
    )

    return p1, p2
