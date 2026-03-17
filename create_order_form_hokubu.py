import re
from pathlib import Path
from typing import Optional

import openpyxl
import pandas as pd


SUPPLIER_NAME = "北部市場販売"

TOKUYOU_SHEET = "特養 (北部市場)"
YUHOUSE_SHEET = "ユーハウス(北部市場)"

# 明細開始・終了
DETAIL_START_ROW = 7
DETAIL_END_ROW = 18  # 7～18 = 12行
ROWS_PER_PAGE = DETAIL_END_ROW - DETAIL_START_ROW + 1

# 納品日表示セル（テンプレ実物に合わせる）
TOKUYOU_DELIVERY_CELL = "J4"
YUHOUSE_DELIVERY_CELL = "I4"

# 施設名
TOKUYOU_LABEL = "介護老人福祉施設　いわと"
YUHOUSE_LABEL = "ユーハウス　いわと"


def _norm(s) -> str:
    if s is None:
        return ""
    return re.sub(r"\s+", " ", str(s).replace("\u3000", " ")).strip()


def _parse_mmdd(value):
    if value is None:
        return None
    s = str(value)
    m = re.search(r"(\d{1,2})/(\d{1,2})", s)
    if not m:
        return None
    mm = int(m.group(1))
    dd = int(m.group(2))
    return pd.Timestamp(year=2000, month=mm, day=dd)


def _sanitize_sheet_title(title: str, existing: set[str]) -> str:
    t = re.sub(r'[:\\/*?\[\]]', "-", str(title)).strip()
    if not t:
        t = "Sheet"
    t = t[:31]
    base = t
    i = 2
    while t in existing:
        suffix = f"_{i}"
        t = (base[:31 - len(suffix)] + suffix)
        i += 1
    return t


def _find_col(df: pd.DataFrame, keywords: list[str]) -> Optional[str]:
    cols = [str(c) for c in df.columns]
    for c in cols:
        if all(k in c for k in keywords):
            return c
    return None


def _copy_sheet(wb, base_ws, title):
    ws = wb.copy_worksheet(base_ws)
    ws.title = _sanitize_sheet_title(title, set(wb.sheetnames))
    return ws


def _clear_detail_rows(ws, is_tokuyou: bool):
    # 7～18行の入力欄をクリア
    for r in range(DETAIL_START_ROW, DETAIL_END_ROW + 1):
        # A:使用日, B/C:品名, D/E/F...数量等
        for c in range(1, ws.max_column + 1):
            # テンプレのラベル行・罫線は残しつつ値だけ消す
            if not isinstance(ws.cell(r, c), openpyxl.cell.cell.MergedCell):
                ws.cell(r, c).value = None

    # 備考欄は残す
    ws["A19"] = "備考欄"


def _format_delivery_date(value: str) -> str:
    """
    '4/18土' → '4月18日'
    '4/18'   → '4月18日'
    """
    if value is None:
        return ""

    s = str(value)

    m = re.search(r"(\d{1,2})/(\d{1,2})", s)
    if not m:
        return s  # 変換できなければそのまま

    mm = int(m.group(1))
    dd = int(m.group(2))

    return f"{mm}月{dd}日"


def _write_delivery_date(ws, delivery_value: str, is_tokuyou: bool):
    cell = TOKUYOU_DELIVERY_CELL if is_tokuyou else YUHOUSE_DELIVERY_CELL

    formatted = _format_delivery_date(delivery_value)

    ws[cell] = f"{formatted}納品分"

def _format_qty_with_unit(qty, unit) -> str:
    """
    数量 + 単位 を1セル表示用に整形
    例: 12 + kg -> '12kg'
        3.0 + 袋 -> '3袋'
    """
    if qty is None:
        return ""

    try:
        q = float(qty)
    except Exception:
        return ""

    if q == 0:
        return ""

    # 12.0 -> 12 にする
    if q.is_integer():
        q_str = str(int(q))
    else:
        q_str = str(q)

    unit_str = "" if unit is None else str(unit).strip()
    return f"{q_str}{unit_str}"
    
def _write_row_tokuyou(ws, row_no: int, use_date: str, food_name: str, qty_res, qty_staff, unit):
    ws.cell(row_no, 1).value = use_date      # A 使用日
    ws.cell(row_no, 2).value = food_name     # B 品名
    ws.cell(row_no, 4).value = qty_res if qty_res != 0 else None   # D 入所者
    ws.cell(row_no, 5).value = qty_staff if qty_staff != 0 else None  # E 職員

    total = (qty_res or 0) + (qty_staff or 0)
    ws.cell(row_no, 6).value = _format_qty_with_unit(total, unit)   # F 合計+単位   


def _write_row_yuhouse(ws, row_no: int, use_date: str, food_name: str, qty_res, unit):
    ws.cell(row_no, 1).value = use_date      # A 使用日
    ws.cell(row_no, 2).value = food_name     # B 品名
    ws.cell(row_no, 4).value = _format_qty_with_unit(qty_res, unit)  # D 入所者+単位


def generate_hokubu_order_workbook(
    kenshu_xlsx_path: str | Path,
    template_xlsm_path: str | Path,
    facility_mode: str,   # "tokuyou" or "yuhouse"
    out_path: str | Path,
) -> Path:
    df = pd.read_excel(kenshu_xlsx_path)
    df = df[df["仕入先"].astype(str) == SUPPLIER_NAME].copy()

    if df.empty:
        raise ValueError("北部市場販売のデータが見つかりません。")

    # 施設列の決定
    tok_res = _find_col(df, ["介護老人福祉施設いわと", "入所者"])
    tok_staff = _find_col(df, ["介護老人福祉施設いわと", "職員"])
    yuhouse_res = _find_col(df, ["ケアハウス", "入"]) or _find_col(df, ["ユーハウス", "入"])

    if facility_mode == "tokuyou":
        qty_cols = [tok_res, tok_staff]
        if not tok_res:
            raise KeyError("特養入所者列が見つかりません。")
        if not tok_staff:
            raise KeyError("特養職員列が見つかりません。")
        sheet_name = TOKUYOU_SHEET
        is_tokuyou = True
    else:
        if not yuhouse_res:
            raise KeyError("ユーハウス入所者列が見つかりません。")
        qty_cols = [yuhouse_res]
        sheet_name = YUHOUSE_SHEET
        is_tokuyou = False

    for col in qty_cols:
        df[col] = pd.to_numeric(df[col], errors="coerce").fillna(0)

    # 数量ゼロ行を落とす
    if is_tokuyou:
        df = df[(df[tok_res] != 0) | (df[tok_staff] != 0)].copy()
    else:
        df = df[df[yuhouse_res] != 0].copy()

    # 並び順：納品日 → 使用日 → 食品名
    df["納品日_dt"] = df["納品日"].apply(_parse_mmdd)
    df["使用日_dt"] = df["使用日"].apply(_parse_mmdd)

    grougroup_cols = ["納品日", "使用日", "食品名", "単位"]
    agg_map = {qty_cols[0]: "sum"}
    if is_tokuyou:
        agg_map[qty_cols[1]] = "sum"

    grouped = (
        df.groupby(group_cols, dropna=False)[list(agg_map.keys())]
        .sum()
        .reset_index()
    )
    grouped["納品日_dt"] = grouped["納品日"].apply(_parse_mmdd)
    grouped["使用日_dt"] = grouped["使用日"].apply(_parse_mmdd)
    grouped = grouped.sort_values(["納品日_dt", "使用日_dt", "食品名"]).reset_index(drop=True)

    wb = openpyxl.load_workbook(template_xlsm_path, keep_vba=True)
    base_ws = wb[sheet_name]

    # ベースシートを残さず、コピーしたものを使う
    created = []
    current_delivery = None
　　current_use_date = None
　　page_no = 0
　　row_in_page = 0
　　ws = None

    for _, rec in grouped.iterrows():
        delivery = str(rec["納品日"])
        use_date = str(rec["使用日"])
        food_name = str(rec["食品名"])
　　　　 unit = rec["単位"]
        
　　　　# 納品日が変わったら新しいページ群へ
        if (
    　　　　current_delivery != delivery
   　　　　 or current_use_date != use_date
  　　　　  or row_in_page >= ROWS_PER_PAGE
　　　　):
    　　　　current_delivery = delivery
   　　　　 current_use_date = use_date
   　　　　 page_no += 1
   　　　　 title = f"{delivery}_{use_date}_{'特養' if is_tokuyou else 'ユーハウス'}_{page_no}"
   　　　　 ws = _copy_sheet(wb, base_ws, title)
    　　　　_clear_detail_rows(ws, is_tokuyou=is_tokuyou)
   　　　　 _write_delivery_date(ws, delivery, is_tokuyou=is_tokuyou)
   　　　　 created.append(ws.title)
   　　　　 row_in_page = 0

        target_row = DETAIL_START_ROW + row_in_page

        if is_tokuyou:
           _write_row_tokuyou(
   　　　　　　 ws,
    　　　　　　target_row,
    　　　　　　use_date,
   　　　　　　 food_name,
   　　　　　　 float(rec[tok_res] or 0),
    　　　　　　float(rec[tok_staff] or 0),
   　　　　　　 unit,
)
        else:
            _write_row_yuhouse(
   　　　　　　　 ws,
   　　　　　　　 target_row,
  　　　　　　　  use_date,
  　　　　　　　  food_name,
  　　　　　　　  float(rec[yuhouse_res] or 0),
   　　　　　　　 unit,
)

        row_in_page += 1

        # 12行使い切ったら次ページ。納品日が同じでも続行
        if row_in_page >= ROWS_PER_PAGE:
            row_in_page = ROWS_PER_PAGE

    # 元テンプレシートは削除
    if base_ws.title in wb.sheetnames:
        wb.remove(base_ws)

    out_path = Path(out_path)
    out_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out_path)
    return out_path


def generate_hokubu_order_forms_both_facilities(
    kenshu_xlsx_path: str | Path,
    template_xlsm_path: str | Path,
    out_dir: str | Path,
    out_prefix: str = "北部市場発注書",
):
    out_dir = Path(out_dir)
    out_dir.mkdir(parents=True, exist_ok=True)

    tokuyou_path = out_dir / f"{out_prefix}_特養.xlsm"
    yuhouse_path = out_dir / f"{out_prefix}_ユーハウス.xlsm"

    p1 = generate_hokubu_order_workbook(
        kenshu_xlsx_path=kenshu_xlsx_path,
        template_xlsm_path=template_xlsm_path,
        facility_mode="tokuyou",
        out_path=tokuyou_path,
    )
    p2 = generate_hokubu_order_workbook(
        kenshu_xlsx_path=kenshu_xlsx_path,
        template_xlsm_path=template_xlsm_path,
        facility_mode="yuhouse",
        out_path=yuhouse_path,
    )
    return p1, p2
