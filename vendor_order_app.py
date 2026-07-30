import io
import re
from datetime import datetime

import pandas as pd
import streamlit as st
from openpyxl.styles import Font, Alignment, Border, Side
from openpyxl.worksheet.page import PageMargins
from openpyxl.utils import get_column_letter


st.set_page_config(
    page_title="業者別発注書作成",
    page_icon="📦",
    layout="wide",
)


# ------------------------------------------------------------
# ゆるかわデザイン
# ------------------------------------------------------------
st.markdown(
    """
    <style>
    .stApp {
        background:
            linear-gradient(90deg, rgba(255,170,130,0.035) 1px, transparent 1px),
            linear-gradient(180deg, rgba(255,170,130,0.035) 1px, transparent 1px),
            linear-gradient(180deg, #fffaf7 0%, #fff7fb 52%, #f7fbff 100%);
        background-size: 28px 28px, 28px 28px, auto;
    }

    .block-container {
        max-width: 980px;
        padding-top: 2rem;
        padding-bottom: 3rem;
    }

    h1 {
        display: inline-block;
        color: #ff7f50 !important;
        background: linear-gradient(180deg, #fff5ed 0%, #ffede2 100%);
        border: 2px solid #ffa56e;
        border-radius: 999px;
        padding: 0.35rem 1.35rem 0.45rem;
        font-size: 2.05rem !important;
        font-weight: 850 !important;
        box-shadow: 0 8px 20px rgba(255,149,99,0.14);
    }

    h2 {
        color: #493f49 !important;
        background: rgba(255,255,255,0.94);
        border: 1px solid #efdcd2;
        border-left: 7px solid #ff9563;
        border-radius: 17px;
        padding: 0.75rem 1rem;
        box-shadow: 0 7px 20px rgba(103,76,89,0.07);
    }

    [data-testid="stFileUploader"] {
        background: rgba(255,255,255,0.94);
        border: 1.5px dashed #e9ae91;
        border-radius: 17px;
        padding: 0.75rem;
        box-shadow: 0 6px 18px rgba(95,75,85,0.06);
    }

    .stButton > button,
    [data-testid="stDownloadButton"] > button {
        width: 100%;
        color: white !important;
        background: linear-gradient(90deg, #ff9b68 0%, #f59ab6 52%, #bca9ef 100%);
        border: none;
        border-radius: 999px;
        font-weight: 820;
        min-height: 2.8rem;
        box-shadow: 0 9px 19px rgba(225,132,145,0.22);
    }

    [data-testid="stAlert"] {
        border-radius: 15px;
    }
    </style>
    """,
    unsafe_allow_html=True,
)


# ------------------------------------------------------------
# 共通処理
# ------------------------------------------------------------
def parse_mmdd(value):
    if value is None or pd.isna(value):
        return None

    match = re.search(r"\d+/\d+", str(value))
    if not match:
        return None

    try:
        return datetime.strptime(match.group(), "%m/%d").replace(year=2000)
    except ValueError:
        return None


def detect_min_usage_date_token(values):
    dates = [parse_mmdd(value) for value in values]
    dates = [value for value in dates if value is not None]

    if not dates:
        return ""

    return min(dates).strftime("%m%d")


def safe_sheet_name(value, used_names):
    name = str(value).strip() or "仕入先未設定"
    name = re.sub(r'[\\/*?:\[\]]', "＿", name)
    name = name[:31] or "仕入先未設定"

    base = name
    number = 2

    while name in used_names:
        suffix = f"_{number}"
        name = f"{base[:31 - len(suffix)]}{suffix}"
        number += 1

    used_names.add(name)
    return name


def apply_order_style(ws):
    thin = Side(style="thin")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    # 見出し行
    for cell in ws[6]:
        cell.font = Font(name="ＭＳ ゴシック", size=12, bold=True)
        cell.alignment = Alignment(
            horizontal="center",
            vertical="center",
            wrap_text=True,
        )
        cell.border = border

    # データ行
    for row in ws.iter_rows(min_row=7, max_row=ws.max_row):
        for cell in row:
            cell.font = Font(name="ＭＳ ゴシック", size=14)
            cell.alignment = Alignment(
                vertical="center",
                wrap_text=False,
            )
            cell.border = border

    # 食品名列
    for cell in ws["B"][6:]:
        cell.alignment = Alignment(
            horizontal="left",
            vertical="center",
            shrink_to_fit=True,
        )

    for row_number in range(1, ws.max_row + 1):
        ws.row_dimensions[row_number].height = 28

    widths = {
        "A": 14,
        "B": 45,
        "C": 13,
        "D": 10,
        "E": 8,
        "F": 8,
        "G": 8,
        "H": 8,
        "I": 8,
        "J": 22,
        "K": 14,
        "L": 14,
    }

    for column, width in widths.items():
        ws.column_dimensions[column].width = width

    ws.page_setup.orientation = "landscape"
    ws.page_setup.paperSize = ws.PAPERSIZE_A4
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.page_margins = PageMargins(
        left=0.3,
        right=0.3,
        top=0.5,
        bottom=0.5,
        header=0.2,
        footer=0.2,
    )
    ws.freeze_panes = "A7"
    ws.print_area = f"A1:L{ws.max_row}"


def create_header(ws, supplier):
    ws.merge_cells("A3:B3")
    ws["A3"] = f"{supplier} 御中"
    ws["A3"].font = Font(name="ＭＳ ゴシック", size=24, bold=True)

    ws["B1"] = "注文書"
    ws["B1"].font = Font(name="ＭＳ ゴシック", size=26, bold=True)
    ws["B1"].alignment = Alignment(horizontal="center")

    ws["K3"] = "(有) ハートミール"
    ws["K3"].font = Font(name="ＭＳ ゴシック", size=20, bold=True)
    ws["K3"].alignment = Alignment(horizontal="right")


def create_orders_from_vendor_sheets(uploaded_file):
    excel_file = pd.ExcelFile(uploaded_file)

    if not excel_file.sheet_names:
        raise ValueError("Excelファイルにシートがありません。")

    required_columns = ["使用日", "食品名", "総合計", "単位"]
    output_data = []
    all_usage_dates = []

    for source_sheet in excel_file.sheet_names:
        df = pd.read_excel(excel_file, sheet_name=source_sheet)
        df = df.dropna(how="all").copy()
        df.columns = [str(column).strip() for column in df.columns]

        missing = [
            column for column in required_columns
            if column not in df.columns
        ]

        if missing:
            raise ValueError(
                f"シート「{source_sheet}」に必要な列がありません: "
                + "、".join(missing)
            )

        supplier = source_sheet

        if "仕入先" in df.columns:
            values = (
                df["仕入先"]
                .dropna()
                .astype(str)
                .str.strip()
            )
            values = values[values != ""]

            if not values.empty:
                supplier = values.iloc[0]

        df["発注数量"] = pd.to_numeric(
            df["総合計"],
            errors="coerce",
        ).fillna(0)

        df = df.loc[df["発注数量"] != 0].copy()

        if df.empty:
            continue

        df["使用日_dt"] = df["使用日"].apply(parse_mmdd)
        df = df.sort_values(
            ["使用日_dt", "食品名"],
            na_position="last",
        )

        df["備考欄"] = (
            df["コメント"].fillna("")
            if "コメント" in df.columns
            else ""
        )

        for column in [
            "鮮度",
            "品温",
            "異物",
            "包装",
            "期限",
            "納品日",
            "検収者",
        ]:
            df[column] = ""

        columns = [
            "使用日",
            "食品名",
            "発注数量",
            "単位",
            "鮮度",
            "品温",
            "異物",
            "包装",
            "期限",
            "備考欄",
            "納品日",
            "検収者",
        ]

        order_df = df[columns].copy()
        order_df["使用日"] = order_df["使用日"].mask(
            order_df["使用日"].duplicated(),
            "",
        )

        all_usage_dates.extend(df["使用日"].dropna().tolist())
        output_data.append((supplier, order_df))

    if not output_data:
        raise ValueError("発注数量が入力されたデータが見つかりません。")

    buffer = io.BytesIO()
    used_names = set()

    with pd.ExcelWriter(buffer, engine="openpyxl") as writer:
        sheet_map = []

        for supplier, order_df in output_data:
            sheet_name = safe_sheet_name(supplier, used_names)

            order_df.to_excel(
                writer,
                sheet_name=sheet_name,
                index=False,
                startrow=5,
            )

            sheet_map.append((sheet_name, supplier))

        workbook = writer.book

        for sheet_name, supplier in sheet_map:
            ws = workbook[sheet_name]
            apply_order_style(ws)
            create_header(ws, supplier)

            ws["C6"] = "発注数量"
            ws["J6"] = "備考欄"

    token = detect_min_usage_date_token(all_usage_dates)
    filename = (
        f"発注書_全業者_{token}.xlsx"
        if token
        else "発注書_全業者.xlsx"
    )

    buffer.seek(0)
    return buffer.read(), filename, len(output_data)


# ------------------------------------------------------------
# 画面
# ------------------------------------------------------------
st.title("業者別発注書作成")
st.caption("🟠 業者別仕訳表をアップロード　🟢 全シートを一括処理　🔵 発注書を自動作成")

st.header("📦 全業者の発注書を一括作成")

st.write(
    "業者別仕訳表にあるすべてのシートを読み込み、"
    "各シートの「総合計」を発注数量として、"
    "業者別の発注書シートを作成します。"
)

uploaded_file = st.file_uploader(
    "業者別仕訳表（複数シートのExcel）をアップロード",
    type=["xlsx"],
    key="vendor_order_source",
)

st.info(
    "各シートに「使用日」「食品名」「総合計」「単位」の列が必要です。"
    "「コメント」列がある場合は備考欄へ引き継ぎます。"
)

if uploaded_file:
    try:
        if st.button("📦 全業者の発注書を作成する"):
            data, filename, sheet_count = create_orders_from_vendor_sheets(
                uploaded_file
            )

            st.session_state["vendor_order_data"] = data
            st.session_state["vendor_order_filename"] = filename

            st.success(
                f"{sheet_count}業者分の発注書を作成しました！"
            )

        if "vendor_order_data" in st.session_state:
            st.download_button(
                "📥 発注書をダウンロード",
                data=st.session_state["vendor_order_data"],
                file_name=st.session_state["vendor_order_filename"],
                mime=(
                    "application/vnd.openxmlformats-officedocument."
                    "spreadsheetml.sheet"
                ),
            )

    except Exception as error:
        st.error("発注書の作成中にエラーが発生しました。")
        st.exception(error)
