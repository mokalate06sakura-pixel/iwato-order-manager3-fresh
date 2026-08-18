import io
import re
import tempfile
from pathlib import Path
from datetime import datetime

import pandas as pd
import streamlit as st
from openpyxl.styles import Font, Alignment, Border, Side
from openpyxl.worksheet.page import PageMargins
from openpyxl.utils import get_column_letter
# 補助機能は、ファイル不足や内部エラーでアプリ全体が停止しないよう安全に読み込む
MARUHACHI_IMPORT_ERROR = None
HOKUBU_IMPORT_ERROR = None

try:
    from create_order_form_maruhachi import generate_maruhachi_order_forms_both_facilities
except Exception as exc:
    generate_maruhachi_order_forms_both_facilities = None
    MARUHACHI_IMPORT_ERROR = exc

try:
    from create_order_form_hokubu import generate_hokubu_order_forms_both_facilities
except Exception as exc:
    generate_hokubu_order_forms_both_facilities = None
    HOKUBU_IMPORT_ERROR = exc
# ------------------------------------------------------------
# Streamlit 基本設定
# ------------------------------------------------------------
st.set_page_config(page_title="発注・検収サポートシステム", layout="wide")

# ------------------------------------------------------------
# 🌸 Streamlit かわいい安定版テーマ
# ------------------------------------------------------------
def apply_cute_theme():
    st.html(
        """
        <style>

        /* ====================================================
           全体
        ==================================================== */
        .stApp {
            background:
                linear-gradient(
                    180deg,
                    #FFF8FC 0%,
                    #F7FAFF 55%,
                    #FFFDF8 100%
                );
        }

        .block-container {
            padding-top: 2.4rem;
            padding-bottom: 3rem;
            max-width: 1080px;
        }

        /* ====================================================
           アプリ上部タイトル
        ==================================================== */
        .app-hero {
            background:
                linear-gradient(
                    135deg,
                    #FFF0F7 0%,
                    #F1F7FF 55%,
                    #FFF9EC 100%
                );
            border: 1px solid #F0DCE9;
            border-radius: 24px;
            padding: 24px 28px;
            margin-bottom: 24px;
            box-shadow:
                0 7px 22px rgba(110, 85, 120, 0.09);
        }

        .app-hero-title {
            font-size: 30px;
            font-weight: 800;
            color: #6B4E71;
            margin-bottom: 8px;
        }

        .app-hero-sub {
            font-size: 16px;
            color: #665B68;
            line-height: 1.8;
            margin-bottom: 14px;
        }

        .cute-pill {
            display: inline-block;
            padding: 7px 12px;
            margin-right: 7px;
            margin-bottom: 5px;
            border-radius: 999px;
            font-size: 14px;
            font-weight: 700;
        }

        .pill-pink {
            background: #FFEAF3;
            color: #A55278;
        }

        .pill-green {
            background: #E9F8EE;
            color: #4B8060;
        }

        .pill-blue {
            background: #EAF3FF;
            color: #4C7199;
        }

        /* ====================================================
           各ページのメインカード
        ==================================================== */
        .main-feature-card {
            background: rgba(255,255,255,0.92);
            border: 1px solid #F0DCE9;
            border-radius: 22px;
            padding: 24px 28px;
            margin-top: 4px;
            margin-bottom: 22px;
            box-shadow:
                0 6px 18px rgba(120,90,120,0.08);
        }

        .main-feature-title {
            font-size: 24px;
            font-weight: 800;
            color: #6B4E71;
            margin-bottom: 10px;
        }

        .main-feature-sub {
            font-size: 16px;
            color: #594F5C;
            line-height: 1.85;
        }

        /* ====================================================
           注意書き
        ==================================================== */
        .main-note-box {
            background: #FFF8E8;
            border-left: 5px solid #F1B66B;
            border-radius: 12px;
            padding: 12px 16px;
            margin-top: 15px;
            color: #6A5643;
            line-height: 1.7;
        }

        .main-note-blue {
            background: #EEF6FF;
            border-left: 5px solid #8BB9E8;
            border-radius: 12px;
            padding: 12px 16px;
            margin-top: 15px;
            color: #4B6075;
            line-height: 1.7;
        }

        .main-note-green {
            background: #EFF9F2;
            border-left: 5px solid #8BC69D;
            border-radius: 12px;
            padding: 12px 16px;
            margin-top: 15px;
            color: #496453;
            line-height: 1.7;
        }

        /* ====================================================
           見出し
        ==================================================== */
        h1 {
            color: #6B4E71;
            font-weight: 800;
        }

        h2, h3 {
            color: #59445E;
            font-weight: 750;
        }

        /* ====================================================
           ボタン
        ==================================================== */
        .stButton > button {
            border-radius: 16px;
            font-weight: 700;
            min-height: 48px;
            border: 1px solid #E7D7E5;
            box-shadow:
                0 3px 9px rgba(100,80,110,0.08);
        }

        .stButton > button:hover {
            border-color: #D9B8D2;
            box-shadow:
                0 5px 12px rgba(100,80,110,0.12);
        }

        /* ====================================================
           ダウンロードボタン
        ==================================================== */
        [data-testid="stDownloadButton"] button {
            border-radius: 16px;
            font-weight: 700;
            min-height: 48px;
            border: 1px solid #D8E5D9;
            box-shadow:
                0 3px 9px rgba(90,120,100,0.08);
        }

        /* ====================================================
           ファイルアップロード
        ==================================================== */
        [data-testid="stFileUploader"] {
            background: #FBFCFF;
            border: 1px solid #E1E8F2;
            border-radius: 18px;
            padding: 8px;
        }

        [data-testid="stFileUploaderDropzone"] {
            border-radius: 14px;
            background: #F5F8FC;
        }

        /* ====================================================
           通知
        ==================================================== */
        [data-testid="stAlert"] {
            border-radius: 14px;
        }

        /* ====================================================
           サイドバー
        ==================================================== */
        [data-testid="stSidebar"] {
            background:
                linear-gradient(
                    180deg,
                    #FFF8FB 0%,
                    #F7F4F1 100%
                );
        }

        .sidebar-section-title {
            margin-bottom: 0.8rem;
            padding: 11px 12px;
            border-radius: 13px;
            background: #F6EAF4;
            color: #6B4E71;
            font-weight: 800;
            font-size: 18px;
        }

        .menu-group-title-blue {
            margin-top: 0.8rem;
            margin-bottom: 0.5rem;
            padding: 8px 10px;
            border-radius: 10px;
            background: #EAF4FF;
            color: #3F6F9F;
            font-weight: 700;
        }

        .menu-group-title-orange {
            margin-top: 1.2rem;
            margin-bottom: 0.5rem;
            padding: 8px 10px;
            border-radius: 10px;
            background: #FFF1E8;
            color: #B85C38;
            font-weight: 700;
        }

        /* ====================================================
           小さいカード
        ==================================================== */
        .mini-card {
            background: rgba(255,255,255,0.90);
            border: 1px solid #E7E2EB;
            border-radius: 17px;
            padding: 15px 17px;
            margin-bottom: 10px;
        }

        .mini-card-title {
            color: #6B4E71;
            font-size: 17px;
            font-weight: 800;
            margin-bottom: 5px;
        }

        .mini-card-sub {
            color: #69616C;
            font-size: 14px;
            line-height: 1.6;
        }

        </style>
        """
    )


apply_cute_theme()

# ------------------------------------------------------------
# 共通ユーティリティ
# ------------------------------------------------------------
def parse_mmdd(value: str):
    """文字列 '12/8月' などから月日だけ抜き出して datetime に変換"""
    if value is None:
        return None
    s = str(value)
    m = re.search(r"\d+/\d+", s)
    if not m:
        return None
    try:
        return datetime.strptime(m.group(), "%m/%d").replace(year=2000)
    except Exception:
        return None


# ------------------------------------------------------------
# ①・② 共通 Excel印刷書式
# A3縦 / 罫線 / 納品日区切り線 / 行高26 / 文字16
# ------------------------------------------------------------
def apply_inspection_print_style(ws):
    # 基本フォント
    body_font = Font(
        name="ＭＳ ゴシック",
        size=16
    )

    header_font = Font(
        name="ＭＳ ゴシック",
        size=16,
        bold=True
    )

    # 罫線
    thin = Side(
        style="thin",
        color="000000"
    )

    medium = Side(
        style="medium",
        color="000000"
    )

    thin_border = Border(
        left=thin,
        right=thin,
        top=thin,
        bottom=thin
    )

    # A3縦
    ws.page_setup.paperSize = ws.PAPERSIZE_A3
    ws.page_setup.orientation = ws.ORIENTATION_PORTRAIT

    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0

    ws.print_options.horizontalCentered = True

    ws.page_margins = PageMargins(
        left=0.25,
        right=0.25,
        top=0.35,
        bottom=0.35,
        header=0.15,
        footer=0.15
    )

    # 全セル
    for row in range(1, ws.max_row + 1):
        ws.row_dimensions[row].height = 26

        for col in range(1, ws.max_column + 1):
            cell = ws.cell(
                row=row,
                column=col
            )

            if row == 1:
                cell.font = header_font
                cell.alignment = Alignment(
                    horizontal="center",
                    vertical="center",
                    wrap_text=True
                )
            else:
                cell.font = body_font
                cell.alignment = Alignment(
                    vertical="center",
                    wrap_text=False
                )

            cell.border = thin_border

    # 見出し名と列番号
    header_map = {}

    for col in range(1, ws.max_column + 1):
        value = ws.cell(
            row=1,
            column=col
        ).value

        if value is not None:
            header_map[str(value).strip()] = col

    # 納品日が変わったら上に太線
    delivery_col = header_map.get("納品日")

    if delivery_col is not None:
        previous_value = None

        for row in range(2, ws.max_row + 1):
            current_value = ws.cell(
                row=row,
                column=delivery_col
            ).value

            current_text = (
                ""
                if current_value is None
                else str(current_value).strip()
            )

            if row == 2:
                previous_value = current_text
                continue

            if current_text != previous_value:
                for col in range(1, ws.max_column + 1):
                    cell = ws.cell(
                        row=row,
                        column=col
                    )

                    cell.border = Border(
                        left=thin,
                        right=thin,
                        top=medium,
                        bottom=thin
                    )

            previous_value = current_text

    # 列幅
    width_map = {
        "納品日": 15,
        "使用日": 15,
        "朝昼夕": 11,
        "仕入先": 24,
        "食品名": 38,
        "換算値": 13,
        "総合計": 13,
        "単位": 11,
        "特養入所者": 15,
        "特養職員": 15,
        "ユーハウス": 15,
        "コメント": 28,
    }

    for header, width in width_map.items():
        col_num = header_map.get(header)

        if col_num is not None:
            letter = get_column_letter(col_num)
            ws.column_dimensions[letter].width = width

    # 食品名は左揃え
    food_col = header_map.get("食品名")

    if food_col is not None:
        for row in range(2, ws.max_row + 1):
            ws.cell(
                row=row,
                column=food_col
            ).alignment = Alignment(
                horizontal="left",
                vertical="center",
                wrap_text=False,
                shrink_to_fit=True
            )

    # 印刷範囲
    ws.print_area = (
        f"A1:"
        f"{get_column_letter(ws.max_column)}"
        f"{ws.max_row}"
    )

    # 各ページに見出し行を表示
    ws.print_title_rows = "1:1"

def detect_min_usage_date_token(df, col="使用日"):
    """使用日の最も古い日付を MMDD 形式 '1208' のように返す"""
    if col not in df.columns:
        return ""

    dt_list = [
        parse_mmdd(v)
        for v in df[col]
    ]

    dt_list = [
        d
        for d in dt_list
        if d is not None
    ]

    if not dt_list:
        return ""

    return min(dt_list).strftime("%m%d")

# ------------------------------------------------------------
# ① 検収簿整形ロジック（修正版：不要列を削除）
# ------------------------------------------------------------
def _is_blank(value):
    """Excelの空白セル（NaN、None、空文字、空白だけの文字列）を判定する。"""
    return pd.isna(value) or (isinstance(value, str) and value.strip() == "")


def apply_ek_blank_rows_and_f_zero(df):
    """VBA「EK空行削除_後にF空白へ0」と同じデータ整形を行う。

    元ファイルの見出し行は pandas が読み取る際に除外されているため、
    ここではデータ行だけを対象にする。
    """
    if df.shape[1] >= 11:
        # ExcelのE～K列（0始まりでは4～10）がすべて空白の行を削除
        ek_all_blank = df.iloc[:, 4:11].apply(
            lambda row: all(_is_blank(value) for value in row), axis=1
        )
        df = df.loc[~ek_all_blank].copy()

    if df.shape[1] >= 6:
        # ExcelのF列（0始まりでは5）の空白を0で埋める
        f_col = df.columns[5]
        f_blank = df[f_col].map(_is_blank)
        df.loc[f_blank, f_col] = 0

    return df


def format_inspection_workbook(uploaded_file):
    df = pd.read_excel(uploaded_file, header=[6, 7])

    # VBA「EK空行削除_後にF空白へ0」を自動適用
    df = apply_ek_blank_rows_and_f_zero(df)

    # ------------------------------------------------------------
    # MultiIndex → フラット化
    # ------------------------------------------------------------
    flat_cols = []

    for top, sub in df.columns:
        top = "" if str(top).startswith("Unnamed") else str(top)
        sub = "" if str(sub).startswith("Unnamed") else str(sub)

        if top == "":
            flat_cols.append(sub)

        elif sub == "":
            flat_cols.append(top)

        else:
            flat_cols.append(f"{top}_{sub}")

    df.columns = flat_cols

    # ------------------------------------------------------------
    # 換算値の空白を0
    # ------------------------------------------------------------
    if "換算値" in df.columns:
        conversion_blank = df["換算値"].map(_is_blank)
        df.loc[conversion_blank, "換算値"] = 0

    # ------------------------------------------------------------
    # 欠損補完
    # ------------------------------------------------------------
    for col in ["納品日", "使用日", "朝昼夕", "仕入先"]:
        if col in df.columns:
            df[col] = df[col].ffill()

    # ------------------------------------------------------------
    # 朝昼夕の並び順
    # ------------------------------------------------------------
    order_map = {
        "朝食": 1,
        "昼食": 2,
        "夕食": 3,
    }

    df["食事順"] = (
        df["朝昼夕"]
        .map(order_map)
        .fillna(0)
    )

    # ------------------------------------------------------------
    # ソート
    # ------------------------------------------------------------
    df = df.sort_values(
        ["使用日", "食事順", "食品名"]
    )

    # ------------------------------------------------------------
    # 特養・ユーハウスの元列を検索
    # ------------------------------------------------------------

    # 特養入所者
    iwato_in = [
        c for c in df.columns
        if (
            "いわと" in str(c)
            and "入所" in str(c)
            and "職員" not in str(c)
        )
    ]

    # 特養職員
    iwato_staff = [
        c for c in df.columns
        if (
            "いわと" in str(c)
            and "職員" in str(c)
        )
    ]

    # ユーハウス
    yuhouse_in = [
        c for c in df.columns
        if (
            (
                "ケアハウス" in str(c)
                or "ユーハウス" in str(c)
                or "ユー" in str(c)
            )
            and (
                "入所者" in str(c)
                or "入居者" in str(c)
                or "入" in str(c)
            )
            and "職員" not in str(c)
        )
    ]

    # ------------------------------------------------------------
    # 必要列
    # ------------------------------------------------------------
    needed_cols = [
        "納品日",
        "使用日",
        "朝昼夕",
        "仕入先",
        "食品名",
        "換算値",
        "総合計",
        "単位",
    ]

    if iwato_in:
        needed_cols.append(iwato_in[0])

    if iwato_staff:
        needed_cols.append(iwato_staff[0])

    if yuhouse_in:
        needed_cols.append(yuhouse_in[0])

    # 実際に存在する列だけ残す
    needed_cols = [
        c for c in needed_cols
        if c in df.columns
    ]

    df_out = df[needed_cols].copy()

    # ------------------------------------------------------------
    # ★列名を統一
    # ------------------------------------------------------------
    rename_map = {}

    if iwato_in:
        rename_map[iwato_in[0]] = "特養入所者"

    if iwato_staff:
        rename_map[iwato_staff[0]] = "特養職員"

    if yuhouse_in:
        rename_map[yuhouse_in[0]] = "ユーハウス"

    df_out = df_out.rename(
        columns=rename_map
    )

    # ------------------------------------------------------------
    # 必須の数量列が取得できたか確認
    # ------------------------------------------------------------
    missing_facility_cols = []

    if "特養入所者" not in df_out.columns:
        missing_facility_cols.append("特養入所者")

    if "特養職員" not in df_out.columns:
        missing_facility_cols.append("特養職員")

    if "ユーハウス" not in df_out.columns:
        missing_facility_cols.append("ユーハウス")

    if missing_facility_cols:
        raise ValueError(
            "検収簿から次の数量列を取得できませんでした："
            + "、".join(missing_facility_cols)
            + "。元の検収記録簿の見出しを確認してください。"
        )

    # ------------------------------------------------------------
    # 換算値を再確認
    # ------------------------------------------------------------
    if "換算値" in df_out.columns:
        conversion_blank = (
            df_out["換算値"]
            .map(_is_blank)
        )

        df_out.loc[
            conversion_blank,
            "換算値"
        ] = 0

    # ------------------------------------------------------------
    # Excel出力
    # ------------------------------------------------------------
    buffer = io.BytesIO()

    with pd.ExcelWriter(
        buffer,
        engine="openpyxl"
    ) as writer:

        df_out.to_excel(
            writer,
            index=False,
            sheet_name="検収簿"
        )

        ws = writer.book["検収簿"]

        apply_inspection_print_style(ws)

    buffer.seek(0)

    token = detect_min_usage_date_token(
        df_out,
        "使用日"
    )

    fname = (
        f"検収簿_加工済_{token}.xlsx"
        if token
        else "検収簿_加工済.xlsx"
    )

    return buffer.read(), fname

# ------------------------------------------------------------
# 業者別仕訳表 作成ロジック
# ------------------------------------------------------------
def _safe_sheet_name(value, used_names):
    """仕入先名をExcelで使用可能な一意のシート名へ変換する。"""
    name = str(value).strip() if not pd.isna(value) else "仕入先未設定"
    name = re.sub(r'[\\/*?:\[\]]', '＿', name)
    name = name[:31] or "仕入先未設定"

    base = name
    index = 2
    while name in used_names:
        suffix = f"_{index}"
        name = f"{base[:31 - len(suffix)]}{suffix}"
        index += 1

    used_names.add(name)
    return name


def create_vendor_journal_workbook(uploaded_file):
    """加工済み検収簿から、仕入先ごとの仕訳表を作成する。"""

    # ------------------------------------------------------------
    # Excel読み込み
    # ------------------------------------------------------------
    df = pd.read_excel(uploaded_file)

    if "仕入先" not in df.columns:
        raise ValueError(
            "『仕入先』列が見つかりません。"
            "検収簿（加工済）を選択してください。"
        )

    # ------------------------------------------------------------
    # 列名を統一
    # ------------------------------------------------------------
    rename_map = {}

    for col in df.columns:
        col_text = str(col)

        if (
            "介護老人福祉施設いわと" in col_text
            and "入所者" in col_text
        ):
            rename_map[col] = "特養入所者"

        elif (
            "介護老人福祉施設いわと" in col_text
            and "職員" in col_text
        ):
            rename_map[col] = "特養職員"

        elif (
            (
                "ケアハウス" in col_text
                or "ユーハウス" in col_text
                or "ユー" in col_text
            )
            and (
                "入所者" in col_text
                or "入居者" in col_text
            )
            and "職員" not in col_text
        ):
            rename_map[col] = "ユーハウス"

    df = df.rename(
        columns=rename_map
    )

    # ------------------------------------------------------------
    # 必須列チェック
    # ------------------------------------------------------------
    required_headers = [
        "単位",
        "特養入所者",
        "特養職員",
        "ユーハウス",
    ]

    missing_headers = [
        name
        for name in required_headers
        if name not in df.columns
    ]

    if missing_headers:
        raise ValueError(
            "必要な列が見つかりません: "
            + "、".join(missing_headers)
            + "。検収簿（加工済）の見出しを確認してください。"
        )

    # ------------------------------------------------------------
    # コメント列を最後へ
    # ------------------------------------------------------------
    if "コメント" in df.columns:
        comment_values = df.pop(
            "コメント"
        )
        df["コメント"] = (
            comment_values
        )
    else:
        df["コメント"] = ""

    # ------------------------------------------------------------
    # 仕入先の空欄処理
    # ------------------------------------------------------------
    df["仕入先"] = (
        df["仕入先"]
        .fillna("仕入先未設定")
        .astype(str)
        .str.strip()
    )

    df.loc[
        df["仕入先"] == "",
        "仕入先"
    ] = "仕入先未設定"

    # ------------------------------------------------------------
    # Excel出力
    # ------------------------------------------------------------
    buffer = io.BytesIO()

    used_names = set()

    with pd.ExcelWriter(
        buffer,
        engine="openpyxl"
    ) as writer:

        for (
            table_index,
            (vendor, vendor_df)
        ) in enumerate(
            df.groupby(
                "仕入先",
                sort=True
            ),
            start=1
        ):

            # ----------------------------------------------------
            # シート名
            # ----------------------------------------------------
            sheet_name = (
                _safe_sheet_name(
                    vendor,
                    used_names
                )
            )

            vendor_df = (
                vendor_df
                .reset_index(
                    drop=True
                )
            )

            # ----------------------------------------------------
            # Excelへ書き込み
            # ----------------------------------------------------
            vendor_df.to_excel(
                writer,
                sheet_name=sheet_name,
                index=False
            )

            ws = writer.book[
                sheet_name
            ]

            # ----------------------------------------------------
            # ①・② 共通の印刷書式を適用
            # ----------------------------------------------------
            apply_inspection_print_style(
                ws
            )

    # ------------------------------------------------------------
    # 出力ファイル名
    # ------------------------------------------------------------
    buffer.seek(0)

    token = (
        detect_min_usage_date_token(
            df,
            "使用日"
        )
    )

    fname = (
        f"業者別仕訳表_{token}.xlsx"
        if token
        else "業者別仕訳表.xlsx"
    )

    return buffer.read(), fname


# ------------------------------------------------------------
# 注文書 書式設定（いわと／ユーハウス共通）
# ------------------------------------------------------------
def apply_order_style(ws, is_tokuyou=False):
    font_body = Font(name="ＭＳ ゴシック", size=18)
    border = Border(
        left=Side("thin"),
        right=Side("thin"),
        top=Side("thin"),
        bottom=Side("thin")
    )

    header_row = 6

    # --- 6行目：ヘッダー行 ---
    for cell in ws[header_row]:
        cell.font = Font(name="ＭＳ ゴシック", size=12, bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = border

    # --- 7行目以降：データ行 ---
    for row in ws.iter_rows(min_row=header_row + 1):
        for c in row:
            c.font = font_body
            c.border = border
            c.alignment = Alignment(
                vertical="center",
                wrap_text=False,      # 折り返しなし
            )

    # --- 行高 ---
    for i in range(1, ws.max_row + 1):
        ws.row_dimensions[i].height = 30

    # ------------------------------------------------------------
    # 列幅設定（注文書仕様）
    # ------------------------------------------------------------

    # A列：使用日
    ws.column_dimensions["A"].width = 15.18

    # B列：食品名（広く）
    ws.column_dimensions["B"].width = 60.09

    # D〜H列：7.73 に変更（数量・単位・確認欄）
    for col in ["D", "E", "F", "G", "H"]:
        ws.column_dimensions[col].width = 7.73

    # C・I・J・K・L・M は 15.18
    for col in ["C", "I", "J", "K", "L", "M"]:
        ws.column_dimensions[col].width = 15.18

    # 特養用マクロの指定
    if is_tokuyou:
        for col in ["I", "L", "M"]:
            ws.column_dimensions[col].width = 7

        # C～E列の外枠を太線にする（内側の罫線は維持）
        thick = Side(style="thick")
        start_row = 6
        end_row = ws.max_row
        if end_row >= start_row:
            for row in range(start_row, end_row + 1):
                ws.cell(row, 3).border = Border(
                    left=thick,
                    right=ws.cell(row, 3).border.right,
                    top=thick if row == start_row else ws.cell(row, 3).border.top,
                    bottom=thick if row == end_row else ws.cell(row, 3).border.bottom,
                )
                ws.cell(row, 5).border = Border(
                    left=ws.cell(row, 5).border.left,
                    right=thick,
                    top=thick if row == start_row else ws.cell(row, 5).border.top,
                    bottom=thick if row == end_row else ws.cell(row, 5).border.bottom,
                )
            for col in range(3, 6):
                ws.cell(start_row, col).border = Border(
                    left=ws.cell(start_row, col).border.left,
                    right=ws.cell(start_row, col).border.right,
                    top=thick,
                    bottom=ws.cell(start_row, col).border.bottom,
                )
                ws.cell(end_row, col).border = Border(
                    left=ws.cell(end_row, col).border.left,
                    right=ws.cell(end_row, col).border.right,
                    top=ws.cell(end_row, col).border.top,
                    bottom=thick,
                )

    # ------------------------------------------------------------
    # B列（食品名）を縮小して全体表示
    # ------------------------------------------------------------
    for row in ws.iter_rows(min_row=7, max_row=ws.max_row, min_col=2, max_col=2):
        for cell in row:
            cell.alignment = Alignment(
                horizontal="left",
                vertical="center",
                wrap_text=False,        # 折り返しなし
                shrink_to_fit=True      # 縮小して全体を表示
            )

    # ------------------------------------------------------------
    # 印刷設定
    # ------------------------------------------------------------
    ws.page_setup.orientation = "landscape"
    ws.page_setup.paperSize = ws.PAPERSIZE_A4
    ws.page_margins = PageMargins(left=0.3, right=0.3, top=0.5, bottom=0.5)

    # 印刷範囲（A〜M列）
    ws.print_area = f"A1:M{ws.max_row}"



# ------------------------------------------------------------
# ヘッダー（いわと）
# ------------------------------------------------------------
def create_header_iwato(ws, supplier):
    ws.merge_cells("A3:B3")
    ws["A3"] = f"{supplier} 御中"
    ws["A3"].font = Font(name="ＭＳ ゴシック", size=28, bold=True)

    ws["B1"] = "注文書（介護老人福祉施設いわと）"
    ws["B1"].alignment = Alignment(horizontal="center")
    ws["B1"].font = Font(name="ＭＳ ゴシック", size=26, bold=True)

    ws["K3"] = "(有) ハートミール"
    ws["K3"].alignment = Alignment(horizontal="right")
    ws["K3"].font = Font(name="ＭＳ ゴシック", size=24, bold=True)



# ------------------------------------------------------------
# ヘッダー（ユーハウス）
# ------------------------------------------------------------
def create_header_yuhouse(ws, supplier):
    ws.merge_cells("A3:B3")
    ws["A3"] = f"{supplier} 御中"
    ws["A3"].font = Font(name="ＭＳ ゴシック", size=28, bold=True)

    ws["B1"] = "注文書（ユーハウスいわと）"
    ws["B1"].alignment = Alignment(horizontal="center")
    ws["B1"].font = Font(name="ＭＳ ゴシック", size=26, bold=True)

    ws["K3"] = "(有) ハートミール"
    ws["K3"].alignment = Alignment(horizontal="right")
    ws["K3"].font = Font(name="ＭＳ ゴシック", size=24, bold=True)


# ------------------------------------------------------------
# ③ 注文書作成
# 特養 / ユーハウス
# ------------------------------------------------------------
def create_order_workbook(uploaded_file, order_type):

    df = pd.read_excel(uploaded_file)

    # ------------------------------------------------------------
    # 基本必須列チェック
    # ------------------------------------------------------------
    required_cols = [
        "使用日",
        "仕入先",
        "食品名",
        "単位",
    ]

    missing_cols = [
        c for c in required_cols
        if c not in df.columns
    ]

    if missing_cols:
        raise ValueError(
            "必要な列が見つかりません："
            + "、".join(missing_cols)
            + "。①検収簿整形で作成した"
              "加工済み検収簿を使用してください。"
        )

    # ------------------------------------------------------------
    # 欠損補完
    # ------------------------------------------------------------
    for c in [
        "使用日",
        "仕入先",
        "食品名",
        "単位",
    ]:
        df[c] = df[c].ffill()

    df["使用日"] = (
        df["使用日"]
        .astype(str)
    )

    # ------------------------------------------------------------
    # 特養
    # ------------------------------------------------------------
    if "特養" in order_type:

        raw_qty = "特養入所者"
        raw_staff = "特養職員"

        missing_tokuyou = []

        if raw_qty not in df.columns:
            missing_tokuyou.append(
                "特養入所者"
            )

        if raw_staff not in df.columns:
            missing_tokuyou.append(
                "特養職員"
            )

        if missing_tokuyou:
            raise ValueError(
                "特養の数量列が見つかりません："
                + "、".join(missing_tokuyou)
                + "。①検収簿整形で作成した"
                  "最新の加工済み検収簿を使用してください。"
            )

        df[raw_qty] = pd.to_numeric(
            df[raw_qty],
            errors="coerce"
        ).fillna(0)

        df[raw_staff] = pd.to_numeric(
            df[raw_staff],
            errors="coerce"
        ).fillna(0)

    # ------------------------------------------------------------
    # ユーハウス
    # ------------------------------------------------------------
    else:

        raw_qty = "ユーハウス"
        raw_staff = None

        if raw_qty not in df.columns:
            raise ValueError(
                "『ユーハウス』列が見つかりません。"
                "①検収簿整形で作成した"
                "最新の加工済み検収簿を使用してください。"
            )

        df[raw_qty] = pd.to_numeric(
            df[raw_qty],
            errors="coerce"
        ).fillna(0)

    # ------------------------------------------------------------
    # 検収用空欄
    # ------------------------------------------------------------
    inspection_cols = [
        "鮮度",
        "品温",
        "異物",
        "包装",
        "期限",
        "備考欄",
        "検収者",
    ]

    for c in inspection_cols:
        if c not in df.columns:
            df[c] = ""

    # 納品日は空欄
    df["納品日"] = ""

    # ------------------------------------------------------------
    # 仕入先整理
    # ------------------------------------------------------------
    df["仕入先"] = (
        df["仕入先"]
        .fillna("")
        .astype(str)
        .str.strip()
    )

    # 仕入先空欄は除外
    df = df[
        df["仕入先"] != ""
    ].copy()

    suppliers = (
        df["仕入先"]
        .drop_duplicates()
        .tolist()
    )

    if not suppliers:
        raise ValueError(
            "仕入先が見つかりません。"
        )

    # ------------------------------------------------------------
    # Excel出力
    # ------------------------------------------------------------
    buffer = io.BytesIO()

    used_sheet_names = set()
    created_sheet_count = 0

    with pd.ExcelWriter(
        buffer,
        engine="openpyxl"
    ) as writer:

        for supplier in suppliers:

            sub = df[
                df["仕入先"] == supplier
            ].copy()

            # ----------------------------------------------------
            # 使用日で並び替え
            # ----------------------------------------------------
            sub["使用日_dt"] = (
                sub["使用日"]
                .apply(parse_mmdd)
            )

            sub = sub.sort_values(
                [
                    "使用日_dt",
                    "食品名",
                ],
                na_position="last",
            )

            # ----------------------------------------------------
            # 特養
            # ----------------------------------------------------
            if "特養" in order_type:

                # 加工済み検収簿では
                # 「特養入所者」「特養職員」を使用
                sub = sub.rename(
                    columns={
                        "特養入所者": "入所者",
                        "特養職員": "職員",
                    }
                )

                qty_label = "入所者"
                staff_label = "職員"

                qty_values = pd.to_numeric(
                    sub[qty_label],
                    errors="coerce"
                ).fillna(0)

                staff_values = pd.to_numeric(
                    sub[staff_label],
                    errors="coerce"
                ).fillna(0)

                # ------------------------------------------------
                # 入所者または職員の
                # どちらかに数量があれば残す
                # ------------------------------------------------
                sub = sub.loc[
                    (qty_values != 0)
                    |
                    (staff_values != 0)
                ].copy()

                if sub.empty:
                    continue

                # 再計算
                qty_values = pd.to_numeric(
                    sub[qty_label],
                    errors="coerce"
                ).fillna(0)

                staff_values = pd.to_numeric(
                    sub[staff_label],
                    errors="coerce"
                ).fillna(0)

                # 0は注文書では空欄
                sub[qty_label] = (
                    sub[qty_label]
                    .astype(object)
                )

                sub[staff_label] = (
                    sub[staff_label]
                    .astype(object)
                )

                sub.loc[
                    qty_values == 0,
                    qty_label
                ] = ""

                sub.loc[
                    staff_values == 0,
                    staff_label
                ] = ""

            # ----------------------------------------------------
            # ユーハウス
            # ----------------------------------------------------
            else:

                # 加工済み検収簿の
                # 「ユーハウス」を使用
                sub = sub.rename(
                    columns={
                        "ユーハウス":
                        "ユーハウス入居者"
                    }
                )

                qty_label = (
                    "ユーハウス入居者"
                )

                staff_label = None

                qty_values = pd.to_numeric(
                    sub[qty_label],
                    errors="coerce"
                ).fillna(0)

                # 数量0は除外
                sub = sub.loc[
                    qty_values != 0
                ].copy()

                if sub.empty:
                    continue

            # ----------------------------------------------------
            # 出力列
            # ----------------------------------------------------
            col_order = [
                "使用日",
                "食品名",
                qty_label,
                "単位",
            ]

            if staff_label:
                col_order.append(
                    staff_label
                )

            col_order += [
                "鮮度",
                "品温",
                "異物",
                "包装",
                "期限",
                "備考欄",
                "納品日",
                "検収者",
            ]

            for c in col_order:
                if c not in sub.columns:
                    sub[c] = ""

            sub = sub[
                col_order
            ].copy()

            # ----------------------------------------------------
            # 同じ使用日は最初だけ表示
            # ----------------------------------------------------
            sub["使用日"] = (
                sub["使用日"].mask(
                    sub["使用日"]
                    .duplicated(),
                    ""
                )
            )

            # ----------------------------------------------------
            # Excelシート名
            # ----------------------------------------------------
            sheet_name = re.sub(
                r'[\\/*?:\[\]]',
                '＿',
                str(supplier)
            )

            sheet_name = (
                sheet_name[:31]
                or "仕入先"
            )

            base_sheet_name = (
                sheet_name
            )

            index = 2

            while (
                sheet_name
                in used_sheet_names
            ):
                suffix = f"_{index}"

                sheet_name = (
                    base_sheet_name[
                        :31 - len(suffix)
                    ]
                    + suffix
                )

                index += 1

            used_sheet_names.add(
                sheet_name
            )

            # ----------------------------------------------------
            # Excelへ書き込み
            # ----------------------------------------------------
            sub.to_excel(
                writer,
                sheet_name=sheet_name,
                index=False,
                startrow=5,
            )

            created_sheet_count += 1

            ws = writer.book[
                sheet_name
            ]

            apply_order_style(
                ws,
                is_tokuyou=(
                    "特養"
                    in order_type
                ),
            )

            # ----------------------------------------------------
            # ヘッダー
            # ----------------------------------------------------
            if "特養" in order_type:

                create_header_iwato(
                    ws,
                    supplier
                )

            else:

                create_header_yuhouse(
                    ws,
                    supplier
                )

                ws["C6"] = (
                    "ユーハウス入居者"
                )

        # --------------------------------------------------------
        # 数量ありのシートが0件
        # --------------------------------------------------------
        if created_sheet_count == 0:
            raise ValueError(
                "注文数量が入力されている"
                "データが見つかりませんでした。"
            )

    # ------------------------------------------------------------
    # ファイル名
    # ------------------------------------------------------------
    token = detect_min_usage_date_token(
        df,
        "使用日"
    )

    if "特養" in order_type:
        base_name = (
            "注文書_いわと"
        )
    else:
        base_name = (
            "注文書_ユーハウス"
        )

    if token:
        fname = (
            f"{base_name}_{token}.xlsx"
        )
    else:
        fname = (
            f"{base_name}.xlsx"
        )

    buffer.seek(0)

    return buffer.read(), fname


# ------------------------------------------------------------
# ------------------------------------------------------------
# 🖥️ UI構築
# ------------------------------------------------------------

# ============================================================
# 🌸 アプリ上部
# ============================================================
st.html(
    '<div class="app-hero">'
    '<div class="app-hero-title">🌷 発注・検収サポートシステム</div>'
    '<div class="app-hero-sub">毎日の発注・検収業務を、やさしく・かんたんに。</div>'
    '<span class="cute-pill pill-pink">🧾 検収簿を整形</span>'
    '<span class="cute-pill pill-green">📦 発注書を自動作成</span>'
    '<span class="cute-pill pill-blue">✨ 業務をかんたんに</span>'
    '</div>',
)

# ============================================================
# サイドバー
# ============================================================
with st.sidebar:
    st.html(
        '<div class="sidebar-section-title">🌸 ユーザーメニュー</div>'
    )

    st.html(
        '<div class="menu-group-title-blue">▼ 📘 検収・注文関連</div>'
    )

    page = st.radio(
        "画面を選択してください",
        [
            "① 検収簿整形",
            "② 業者別仕訳表",
            "③ 注文書作成",
            "④ 丸八発注書作成",
            "⑤ 北部市場発注書作成",
        ],
        key="main_page_selector",
        label_visibility="collapsed",
    )

    st.html(
        '<div class="menu-group-title-orange">'
        '📦 ④・⑤は業者別発注書メニューです'
        '</div>'
    )


# ============================================================
# ① 検収簿整形
# ============================================================
if page == "① 検収簿整形":

    st.html(
        '<div class="main-feature-card">'
        '<div class="main-feature-title">🌷 ① 検収簿を整える</div>'
        '<div class="main-feature-sub">'
        '献ダテマンから出力した検収記録簿を、<br>'
        '発注・検収アプリで使いやすい形に整えます。'
        '</div>'
        '<div class="main-note-box">'
        '📌 献ダテマンから出力したファイルを<br>'
        '<b>「検収記録簿_原本.xlsx」</b> の名前で保存して下さい。'
        '</div>'
        '</div>',
    )

    ins_file = st.file_uploader(
        "📄 検収簿（原本 Excel）をアップロード",
        type=["xlsx"],
        key="ins",
    )

    if ins_file is not None:
        if st.button(
            "📘 検収簿を整形する",
            key="btn_ins",
            use_container_width=True,
        ):
            try:
                (
                    st.session_state["ins_data"],
                    st.session_state["ins_fname"],
                ) = format_inspection_workbook(ins_file)
                st.success("🌸 検収簿の整形が完了しました！")
            except Exception as e:
                st.error("検収簿の整形中にエラーが発生しました。")
                st.exception(e)

        if (
            "ins_data" in st.session_state
            and "ins_fname" in st.session_state
        ):
            st.download_button(
                label="📥 検収簿（加工済）をダウンロード",
                data=st.session_state["ins_data"],
                file_name=st.session_state["ins_fname"],
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                key="download_inspection",
                use_container_width=True,
            )


# ============================================================
# ② 業者別仕訳表
# ============================================================
elif page == "② 業者別仕訳表":

    st.html(
        '<div class="main-feature-card">'
        '<div class="main-feature-title">📊 ② 業者別仕訳表を作成</div>'
        '<div class="main-feature-sub">'
        '加工済み検収簿を、仕入先ごとのシートに分けて自動作成します。'
        '</div>'
        '<div class="main-note-blue">'
        '💡 「① 検収簿整形」で作成した <b>検収簿_加工済</b> を使用してください。<br>'
        'A3縦・文字16・行高26・罫線付きで出力します。'
        '</div>'
        '</div>',
    )

    vendor_file = st.file_uploader(
        "📄 検収簿（加工済 Excel）をアップロード",
        type=["xlsx"],
        key="vendor_journal_src",
    )

    if vendor_file is not None:
        try:
            if st.button(
                "📊 業者別仕訳表を作成する",
                key="btn_vendor_journal",
                use_container_width=True,
            ):
                (
                    st.session_state["vendor_journal_data"],
                    st.session_state["vendor_journal_fname"],
                ) = create_vendor_journal_workbook(vendor_file)
                st.success("🌸 業者別仕訳表の作成が完了しました！")

            if (
                "vendor_journal_data" in st.session_state
                and "vendor_journal_fname" in st.session_state
            ):
                st.download_button(
                    label="📥 業者別仕訳表をダウンロード",
                    data=st.session_state["vendor_journal_data"],
                    file_name=st.session_state["vendor_journal_fname"],
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    key="download_vendor_journal",
                    use_container_width=True,
                )
        except Exception as e:
            st.error("業者別仕訳表の作成中にエラーが発生しました。")
            st.exception(e)


# ============================================================
# ③ 注文書作成
# ============================================================
elif page == "③ 注文書作成":

    st.html(
        '<div class="main-feature-card">'
        '<div class="main-feature-title">📗 ③ 注文書を作成</div>'
        '<div class="main-feature-sub">'
        '特養またはユーハウスを選択して、<br>'
        '仕入先別の注文書を自動作成します。'
        '</div>'
        '<div class="main-note-green">'
        '🌿 特養では <b>「入所者」と「職員」</b> を同じ注文書にまとめて出力します。'
        '</div>'
        '</div>',
    )

    order_type = st.radio(
        "作成する注文書の種類を選んでください",
        (
            "特養（介護老人福祉施設いわと）",
            "ユーハウスいわと",
        ),
        horizontal=True,
        key="order_type",
    )

    order_file = st.file_uploader(
        "📄 注文書のもとになる検収簿 Excel",
        type=["xlsx"],
        key="order_src",
    )

    st.info(
        "① 検収簿整形で作成した「検収簿_加工済」ファイルを使用してください。"
    )

    if order_file is not None:
        current_file_id = (
            order_file.name,
            order_file.size,
            order_type,
        )

        previous_file_id = st.session_state.get("order_current_file_id")

        if previous_file_id != current_file_id:
            st.session_state["order_current_file_id"] = current_file_id
            st.session_state.pop("order_data", None)
            st.session_state.pop("order_fname", None)

        if st.button(
            "📗 注文書を作成する",
            key="btn_order",
            use_container_width=True,
        ):
            try:
                order_bytes, order_fname = create_order_workbook(
                    order_file,
                    order_type,
                )
                st.session_state["order_data"] = order_bytes
                st.session_state["order_fname"] = order_fname
                st.success(f"🌸 {order_type} の注文書を作成しました！")
            except Exception as e:
                st.session_state.pop("order_data", None)
                st.session_state.pop("order_fname", None)
                st.error("注文書作成中にエラーが発生しました。")
                st.exception(e)

        if (
            "order_data" in st.session_state
            and "order_fname" in st.session_state
        ):
            st.download_button(
                label="📥 注文書ファイルをダウンロード",
                data=st.session_state["order_data"],
                file_name=st.session_state["order_fname"],
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                key="download_order",
                use_container_width=True,
            )


# ============================================================
# ④ 丸八発注書作成
# ============================================================
elif page == "④ 丸八発注書作成":

    st.html(
        '<div class="main-feature-card">'
        '<div class="main-feature-title">📦 ④ 丸八発注書を作成</div>'
        '<div class="main-feature-sub">'
        '検収簿_加工済・丸八発注書テンプレート・<br>'
        '丸八コード一覧から、特養用とユーハウス用の発注書を同時に作成します。'
        '</div>'
        '<div class="main-note-box">📌 下の3つのファイルをすべて選択してください。</div>'
        '</div>',
    )

    mcol1, mcol2, mcol3 = st.columns(3)

    with mcol1:
        st.html(
            '<div class="mini-card">'
            '<div class="mini-card-title">📄 検収簿_加工済</div>'
            '<div class="mini-card-sub">丸八ヒロタのデータを含む加工済ファイル</div>'
            '</div>',
        )
        kenshu_file = st.file_uploader(
            "検収簿_加工済（xlsx）",
            type=["xlsx"],
            key="kenshu_maruhachi",
        )

    with mcol2:
        st.html(
            '<div class="mini-card">'
            '<div class="mini-card-title">🧾 丸八発注書テンプレ</div>'
            '<div class="mini-card-sub">丸八ヒロタ専用の発注書テンプレート</div>'
            '</div>',
        )
        template_file = st.file_uploader(
            "丸八発注書テンプレ（xlsm）",
            type=["xlsm"],
            key="tpl_maruhachi",
        )

    with mcol3:
        st.html(
            '<div class="mini-card">'
            '<div class="mini-card-title">🏷️ 丸八コード一覧</div>'
            '<div class="mini-card-sub">タグシート付きのコード対応表</div>'
            '</div>',
        )
        tag_file = st.file_uploader(
            "丸八コード一覧（xlsm）",
            type=["xlsm"],
            key="tag_maruhachi",
        )

    btn = st.button(
        "📦 丸八発注書を作成",
        key="btn_maruhachi",
        use_container_width=True,
    )

    if btn:
        if generate_maruhachi_order_forms_both_facilities is None:
            st.error("丸八発注書機能を読み込めませんでした。")
            if MARUHACHI_IMPORT_ERROR is not None:
                st.exception(MARUHACHI_IMPORT_ERROR)

        elif not (kenshu_file and template_file and tag_file):
            st.warning("⚠ 3つのファイルをすべて選択してください。")

        else:
            try:
                with st.spinner("丸八発注書を作成しています…"):
                    with tempfile.TemporaryDirectory() as td:
                        td = Path(td)

                        k_path = td / "kenshu.xlsx"
                        t_path = td / "template.xlsm"
                        m_path = td / "tag.xlsm"

                        k_path.write_bytes(kenshu_file.getbuffer())
                        t_path.write_bytes(template_file.getbuffer())
                        m_path.write_bytes(tag_file.getbuffer())

                        out_dir = td / "out"

                        tokuyou_xlsm, yuhouse_xlsm = (
                            generate_maruhachi_order_forms_both_facilities(
                                kenshu_xlsx_path=k_path,
                                template_xlsm_path=t_path,
                                tag_xlsm_path=m_path,
                                out_dir=out_dir,
                                out_prefix="丸八発注書",
                            )
                        )

                        st.session_state["maruhachi_tokuyou_data"] = tokuyou_xlsm.read_bytes()
                        st.session_state["maruhachi_tokuyou_fname"] = tokuyou_xlsm.name
                        st.session_state["maruhachi_yuhouse_data"] = yuhouse_xlsm.read_bytes()
                        st.session_state["maruhachi_yuhouse_fname"] = yuhouse_xlsm.name

                st.success("🌸 丸八発注書を作成しました！")

            except Exception as e:
                st.session_state.pop("maruhachi_tokuyou_data", None)
                st.session_state.pop("maruhachi_tokuyou_fname", None)
                st.session_state.pop("maruhachi_yuhouse_data", None)
                st.session_state.pop("maruhachi_yuhouse_fname", None)
                st.error("丸八発注書の作成中にエラーが発生しました。")
                st.exception(e)

    if (
        "maruhachi_tokuyou_data" in st.session_state
        and "maruhachi_yuhouse_data" in st.session_state
    ):
        st.markdown("### 📥 作成済みファイル")

        dcol1, dcol2 = st.columns(2)

        with dcol1:
            st.download_button(
                "📥 特養：丸八発注書",
                data=st.session_state["maruhachi_tokuyou_data"],
                file_name=st.session_state["maruhachi_tokuyou_fname"],
                mime="application/vnd.ms-excel.sheet.macroEnabled.12",
                key="download_maruhachi_tokuyou",
                use_container_width=True,
            )

        with dcol2:
            st.download_button(
                "📥 ユーハウス：丸八発注書",
                data=st.session_state["maruhachi_yuhouse_data"],
                file_name=st.session_state["maruhachi_yuhouse_fname"],
                mime="application/vnd.ms-excel.sheet.macroEnabled.12",
                key="download_maruhachi_yuhouse",
                use_container_width=True,
            )


# ============================================================
# ⑤ 北部市場発注書作成
# ============================================================
elif page == "⑤ 北部市場発注書作成":

    st.html(
        '<div class="main-feature-card">'
        '<div class="main-feature-title">🥕 ⑤ 北部市場発注書を作成</div>'
        '<div class="main-feature-sub">'
        '検収簿_加工済と北部市場発注書テンプレートを使って、<br>'
        '特養用・ユーハウス用の発注書を同時に作成します。'
        '</div>'
        '<div class="main-note-green">'
        '🌿 2つのファイルを選択して、「北部市場発注書を作成」を押してください。'
        '</div>'
        '</div>',
    )

    hcol1, hcol2 = st.columns(2)

    with hcol1:
        st.html(
            '<div class="mini-card">'
            '<div class="mini-card-title">📄 検収簿_加工済</div>'
            '<div class="mini-card-sub">北部市場販売のデータを含む加工済ファイル</div>'
            '</div>',
        )
        hokubu_kenshu = st.file_uploader(
            "検収簿_加工済（xlsx）",
            type=["xlsx"],
            key="hokubu_kenshu",
        )

    with hcol2:
        st.html(
            '<div class="mini-card">'
            '<div class="mini-card-title">🧾 北部市場テンプレート</div>'
            '<div class="mini-card-sub">北部市場専用の発注書テンプレート</div>'
            '</div>',
        )
        hokubu_template = st.file_uploader(
            "北部市場発注書テンプレート（xlsm）",
            type=["xlsm"],
            key="hokubu_tpl",
        )

    btn_hokubu = st.button(
        "🥕 北部市場発注書を作成",
        key="btn_hokubu",
        use_container_width=True,
    )

    if btn_hokubu:
        if generate_hokubu_order_forms_both_facilities is None:
            st.error("北部市場発注書機能を読み込めませんでした。")
            if HOKUBU_IMPORT_ERROR is not None:
                st.exception(HOKUBU_IMPORT_ERROR)

        elif not (hokubu_kenshu and hokubu_template):
            st.warning(
                "⚠ 検収簿_加工済 と 北部市場テンプレートを両方選択してください。"
            )

        else:
            try:
                with st.spinner("北部市場発注書を作成しています…"):
                    with tempfile.TemporaryDirectory() as td:
                        td = Path(td)

                        k_path = td / "kenshu.xlsx"
                        t_path = td / "template.xlsm"

                        k_path.write_bytes(hokubu_kenshu.getbuffer())
                        t_path.write_bytes(hokubu_template.getbuffer())

                        out_dir = td / "out"

                        tokuyou_xlsm, yuhouse_xlsm = (
                            generate_hokubu_order_forms_both_facilities(
                                kenshu_xlsx_path=k_path,
                                template_xlsm_path=t_path,
                                out_dir=out_dir,
                                out_prefix="北部市場発注書",
                            )
                        )

                        st.session_state["hokubu_tokuyou_data"] = tokuyou_xlsm.read_bytes()
                        st.session_state["hokubu_tokuyou_fname"] = tokuyou_xlsm.name
                        st.session_state["hokubu_yuhouse_data"] = yuhouse_xlsm.read_bytes()
                        st.session_state["hokubu_yuhouse_fname"] = yuhouse_xlsm.name

                st.success("🌸 北部市場発注書を作成しました！")

            except Exception as e:
                st.session_state.pop("hokubu_tokuyou_data", None)
                st.session_state.pop("hokubu_tokuyou_fname", None)
                st.session_state.pop("hokubu_yuhouse_data", None)
                st.session_state.pop("hokubu_yuhouse_fname", None)
                st.error("北部市場発注書の作成中にエラーが発生しました。")
                st.exception(e)

    if (
        "hokubu_tokuyou_data" in st.session_state
        and "hokubu_yuhouse_data" in st.session_state
    ):
        st.markdown("### 📥 作成済みファイル")

        c1, c2 = st.columns(2)

        with c1:
            st.download_button(
                label="📥 特養：北部市場発注書",
                data=st.session_state["hokubu_tokuyou_data"],
                file_name=st.session_state["hokubu_tokuyou_fname"],
                mime="application/vnd.ms-excel.sheet.macroEnabled.12",
                key="download_hokubu_tokuyou",
                use_container_width=True,
            )

        with c2:
            st.download_button(
                label="📥 ユーハウス：北部市場発注書",
                data=st.session_state["hokubu_yuhouse_data"],
                file_name=st.session_state["hokubu_yuhouse_fname"],
                mime="application/vnd.ms-excel.sheet.macroEnabled.12",
                key="download_hokubu_yuhouse",
                use_container_width=True,
            )
